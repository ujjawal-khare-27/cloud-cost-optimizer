import unittest
import os
import tempfile
from openpyxl import load_workbook

from src.core.aws.excel_report_generator import ExcelReportGenerator


class TestExcelReportGenerator(unittest.TestCase):
    def setUp(self):
        self.report_generator = ExcelReportGenerator()
        self.sample_unused_resources = [
            {
                "ebs": {
                    "unused_ebs_volumes": [
                        {
                            "VolumeId": "vol-0123456789abcdef0",
                            "Size": 8,
                            "State": "available",
                            "AvailabilityZone": "us-east-1a",
                            "CreateTime": "2023-10-01T12:34:56.000Z",
                        },
                        {
                            "VolumeId": "vol-0abcdef1234567890",
                            "Size": 20,
                            "State": "available",
                            "AvailabilityZone": "us-east-1b",
                            "CreateTime": "2023-11-15T08:22:10.000Z",
                        },
                    ]
                }
            },
            {
                "lb": {
                    "no_targets_lb": [
                        {
                            "LoadBalancerName": "test-lb-no-targets",
                            "DNSName": "test-lb-no-targets-1234567890.us-east-1.elb.amazonaws.com",
                            "CreatedTime": "2023-10-01T12:00:00.000Z",
                        }
                    ],
                    "all_unhealthy": [
                        {
                            "LoadBalancerName": "test-lb-all-unhealthy",
                            "DNSName": "test-lb-all-unhealthy-1234567890.us-east-1.elb.amazonaws.com",
                            "CreatedTime": "2023-10-01T12:00:00.000Z",
                        }
                    ],
                }
            },
            {
                "rds": {
                    "rds_with_no_connections": [
                        {
                            "DBInstanceIdentifier": "test-db-instance-1",
                            "DBInstanceClass": "db.t3.micro",
                            "Engine": "mysql",
                            "AllocatedStorage": 20,
                        }
                    ],
                    "rds_instances_with_no_connections": [
                        {
                            "DBInstanceIdentifier": "test-db-instance-1",
                            "DBInstanceClass": "db.t3.micro",
                            "Engine": "mysql",
                            "AllocatedStorage": 20,
                        }
                    ],
                }
            },
        ]

    def tearDown(self):
        # Clean up any temporary files
        if hasattr(self, "temp_file") and os.path.exists(self.temp_file):
            os.unlink(self.temp_file)

    def test_generate_report_creates_file(self):
        """Test that generate_report creates an Excel file."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            self.temp_file = tmp_file.name

        report_path = self.report_generator.generate_report(
            unused_resources=self.sample_unused_resources, region="us-east-1", output_path=self.temp_file
        )

        self.assertEqual(report_path, self.temp_file)
        self.assertTrue(os.path.exists(report_path))

    def test_generate_report_with_default_path(self):
        """Test that generate_report creates a file with default naming."""
        report_path = self.report_generator.generate_report(
            unused_resources=self.sample_unused_resources, region="us-east-1"
        )

        self.assertTrue(report_path.__contains__("aws_unused_resources_report_us-east-1_"))
        self.assertTrue(report_path.endswith(".xlsx"))
        self.assertTrue(os.path.exists(report_path))

        # Clean up
        if os.path.exists(report_path):
            os.unlink(report_path)

    def test_excel_file_structure(self):
        """Test that the generated Excel file has the correct structure."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            self.temp_file = tmp_file.name

        self.report_generator.generate_report(
            unused_resources=self.sample_unused_resources, region="us-east-1", output_path=self.temp_file
        )

        # Load the workbook and verify structure
        workbook = load_workbook(self.temp_file)

        # Check that we have the expected sheets
        expected_sheets = ["Summary", "EBS", "LB", "RDS"]
        self.assertEqual(set(workbook.sheetnames), set(expected_sheets))

        # Check Summary sheet content
        summary_sheet = workbook["Summary"]
        self.assertEqual(summary_sheet["A1"].value, "AWS Unused Resources Report - us-east-1")
        self.assertIn("Generated on:", str(summary_sheet["A2"].value))

        # Check that summary has the expected headers
        self.assertEqual(summary_sheet["A4"].value, "Service")
        self.assertEqual(summary_sheet["B4"].value, "Resource Type")
        self.assertEqual(summary_sheet["C4"].value, "Count")
        self.assertEqual(summary_sheet["D4"].value, "Details")

        workbook.close()

    def test_ebs_sheet_content(self):
        """Test that the EBS sheet contains the correct data."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            self.temp_file = tmp_file.name

        self.report_generator.generate_report(
            unused_resources=self.sample_unused_resources, region="us-east-1", output_path=self.temp_file
        )

        workbook = load_workbook(self.temp_file)
        ebs_sheet = workbook["EBS"]

        # Check title
        self.assertEqual(ebs_sheet["A1"].value, "EBS Unused Resources")

        # Check that we have the expected volume data
        # The first volume should be in row 5 (after title, spacing, and headers)
        self.assertEqual(ebs_sheet["A5"].value, "vol-0123456789abcdef0")
        self.assertEqual(ebs_sheet["B5"].value, 8)
        self.assertEqual(ebs_sheet["C5"].value, "available")
        self.assertEqual(ebs_sheet["D5"].value, "us-east-1a")

        workbook.close()

    def test_lb_sheet_content(self):
        """Test that the LB sheet contains the correct data."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            self.temp_file = tmp_file.name

        self.report_generator.generate_report(
            unused_resources=self.sample_unused_resources, region="us-east-1", output_path=self.temp_file
        )

        workbook = load_workbook(self.temp_file)
        lb_sheet = workbook["LB"]

        # Check title
        self.assertEqual(lb_sheet["A1"].value, "LB Unused Resources")

        # Check that we have both resource types
        self.assertIn("No Targets Lb", [cell.value for cell in lb_sheet["A"]])
        self.assertIn("All Unhealthy", [cell.value for cell in lb_sheet["A"]])

        workbook.close()

    def test_rds_sheet_content(self):
        """Test that the RDS sheet contains the correct data."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            self.temp_file = tmp_file.name

        self.report_generator.generate_report(
            unused_resources=self.sample_unused_resources, region="us-east-1", output_path=self.temp_file
        )

        workbook = load_workbook(self.temp_file)
        rds_sheet = workbook["RDS"]

        # Check title
        self.assertEqual(rds_sheet["A1"].value, "RDS Unused Resources")

        # Check that we have both resource types
        self.assertIn("Rds With No Connections", [cell.value for cell in rds_sheet["A"]])
        self.assertIn("Rds Instances With No Connections", [cell.value for cell in rds_sheet["A"]])

        workbook.close()

    def test_empty_resources_handling(self):
        """Test that empty resources are handled correctly."""
        empty_resources = [{"ebs": {"unused_ebs_volumes": []}}]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            self.temp_file = tmp_file.name

        self.report_generator.generate_report(
            unused_resources=empty_resources, region="us-east-1", output_path=self.temp_file
        )

        workbook = load_workbook(self.temp_file)

        # Should only have Summary sheet when no resources are found
        self.assertEqual(workbook.sheetnames, ["Summary"])

        # Summary should show 0 total resources
        summary_sheet = workbook["Summary"]
        # Find the total row
        total_row = None
        for row in range(1, summary_sheet.max_row + 1):
            if summary_sheet[f"A{row}"].value == "TOTAL":
                total_row = row
                break

        self.assertIsNotNone(total_row)
        self.assertEqual(summary_sheet[f"C{total_row}"].value, 0)

        workbook.close()

    def test_summary_sheet_calculations(self):
        """Test that the summary sheet correctly calculates totals."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            self.temp_file = tmp_file.name

        self.report_generator.generate_report(
            unused_resources=self.sample_unused_resources, region="us-east-1", output_path=self.temp_file
        )

        workbook = load_workbook(self.temp_file)
        summary_sheet = workbook["Summary"]

        # Find the total row
        total_row = None
        for row in range(1, summary_sheet.max_row + 1):
            if summary_sheet[f"A{row}"].value == "TOTAL":
                total_row = row
                break

        self.assertIsNotNone(total_row)
        # Should have 6 total resources: 2 EBS volumes + 2 LB + 2 RDS (both RDS categories have the same instance)
        expected_total = 6  # 2 EBS + 2 LB + 2 RDS (both categories count separately)
        self.assertEqual(summary_sheet[f"C{total_row}"].value, expected_total)

        workbook.close()

    def test_workbook_cleanup(self):
        """Test that the workbook is properly cleaned up."""
        generator = ExcelReportGenerator()

        # Generate a report
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            temp_file = tmp_file.name

        report_path = generator.generate_report(
            unused_resources=self.sample_unused_resources, region="us-east-1", output_path=temp_file
        )

        # Verify file was created
        print(f"Report path: {report_path}")
        self.assertTrue(os.path.exists(report_path))

        # Clean up
        if os.path.exists(report_path):
            os.unlink(report_path)

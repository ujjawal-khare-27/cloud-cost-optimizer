import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:
    """
    Generates Excel reports for AWS unused resources.
    Each service gets its own worksheet with formatted data.
    """

    def __init__(self):
        self.workbook = Workbook()
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")

    def _create_summary_sheet(self, unused_resources: List[Dict[str, Any]], region: str) -> None:
        """Create a summary sheet with overview of all services."""
        # Remove default sheet and create summary
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

        summary_sheet = self.workbook.create_sheet("Summary", 0)

        # Add title
        summary_sheet["A1"] = f"AWS Unused Resources Report - {region}"
        summary_sheet["A1"].font = Font(bold=True, size=16)
        summary_sheet["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        summary_sheet["A2"].font = Font(italic=True)

        # Add summary table headers
        summary_sheet["A4"] = "Service"
        summary_sheet["B4"] = "Resource Type"
        summary_sheet["C4"] = "Count"
        summary_sheet["D4"] = "Details"

        # Style headers
        for col in ["A4", "B4", "C4", "D4"]:
            summary_sheet[col].font = self.header_font
            summary_sheet[col].fill = self.header_fill
            summary_sheet[col].alignment = self.center_alignment
            summary_sheet[col].border = self.border

        row = 5
        total_resources = 0

        for service_data in unused_resources:
            for service_name, resources in service_data.items():
                if isinstance(resources, dict):
                    for resource_type, resource_list in resources.items():
                        if isinstance(resource_list, list) and resource_list:
                            count = len(resource_list)
                            total_resources += count

                            summary_sheet[f"A{row}"] = service_name.upper()
                            summary_sheet[f"B{row}"] = resource_type.replace("_", " ").title()
                            summary_sheet[f"C{row}"] = count
                            summary_sheet[f"D{row}"] = f"See {service_name.upper()} sheet for details"

                            # Style data rows
                            for col in ["A", "B", "C", "D"]:
                                summary_sheet[f"{col}{row}"].border = self.border
                                if col == "C":
                                    summary_sheet[f"{col}{row}"].alignment = self.center_alignment

                            row += 1

        # Add total row (always add it, even if 0)
        summary_sheet[f"A{row}"] = "TOTAL"
        summary_sheet[f"A{row}"].font = Font(bold=True)
        summary_sheet[f"C{row}"] = total_resources
        summary_sheet[f"C{row}"].font = Font(bold=True)
        summary_sheet[f"C{row}"].alignment = self.center_alignment

        for col in ["A", "B", "C", "D"]:
            summary_sheet[f"{col}{row}"].border = self.border

        # Auto-adjust column widths
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_sheet.column_dimensions[column_letter].width = adjusted_width

    def _create_service_sheet(self, service_name: str, resources: Dict[str, Any]) -> None:
        """Create a detailed worksheet for a specific service."""
        sheet_name = service_name.upper()
        worksheet = self.workbook.create_sheet(sheet_name)

        # Add title
        worksheet["A1"] = f"{service_name.upper()} Unused Resources"
        worksheet["A1"].font = Font(bold=True, size=14)

        row = 3

        for resource_type, resource_list in resources.items():
            if isinstance(resource_list, list) and resource_list:
                # Add resource type header
                worksheet[f"A{row}"] = f"{resource_type.replace('_', ' ').title()}"
                worksheet[f"A{row}"].font = Font(bold=True, size=12)
                worksheet[f"A{row}"].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                row += 1

                # Add column headers
                if resource_list:
                    headers = list(resource_list[0].keys())
                    for col_idx, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=row, column=col_idx, value=header.replace("_", " ").title())
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = self.center_alignment
                        cell.border = self.border
                    row += 1

                    # Add data rows
                    for resource in resource_list:
                        for col_idx, header in enumerate(headers, 1):
                            value = resource.get(header, "")
                            cell = worksheet.cell(row=row, column=col_idx, value=str(value))
                            cell.border = self.border
                            if isinstance(value, (int, float)):
                                cell.alignment = self.center_alignment
                        row += 1

                row += 1  # Add spacing between resource types

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def generate_report(self, unused_resources: List[Dict[str, Any]], region: str, output_path: str = None) -> str:
        """
        Generate Excel report for unused resources.

        Args:
            unused_resources: List of dictionaries containing unused resources by service
            region: AWS region
            output_path: Optional custom output path

        Returns:
            Path to the generated Excel file
        """
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            cwd = os.getcwd()
            report_path = f"{cwd.split('cloud-cost-optimizer')[0]}cloud-cost-optimizer/reports/aws_unused_resources_report_{region}_{timestamp}.xlsx"
            output_path = report_path

        # Create summary sheet
        self._create_summary_sheet(unused_resources, region)

        # Create individual service sheets
        for service_data in unused_resources:
            for service_name, resources in service_data.items():
                if isinstance(resources, dict) and any(
                    isinstance(resource_list, list) and resource_list for resource_list in resources.values()
                ):
                    self._create_service_sheet(service_name, resources)

        # Save the workbook
        self.workbook.save(output_path)
        return output_path

    def __del__(self):
        """Clean up workbook when object is destroyed."""
        if hasattr(self, "workbook"):
            self.workbook.close()
